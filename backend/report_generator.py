import pandas as pd
from datetime import datetime
from pathlib import Path
import json
from typing import List, Dict, Any, Optional, Tuple
import asyncio
import base64
import io

class ReportGenerator:
    def __init__(self):
        self.class_names = {0: "Krater", 1: "Tanecik", 2: "Pinhol"}
        self.class_colors = {
            "Krater": "#FF6B6B",    # Red
            "Tanecik": "#4ECDC4",   # Teal
            "Pinhol": "#45B7D1"     # Blue
        }

    # ---------------------------
    # ÜST SEVİYE: Toplu rapor üret
    # ---------------------------
    async def generate_reports(
        self,
        results_data: Optional[List[Dict[str, Any]]] = None,
        base_name: str = "Analiz_Sonuclari",
        out_root: str = "downloads"
    ) -> Dict[str, Any]:
        """
        Excel + JSON raporlarını 'downloads/<base_name>_<timestamp>/reports' altına üretir.
        results_data verilmezse sırasıyla:
          - results/analysis_results.json
          - downloads/latest_results.json
        dosyalarını okumayı dener. Hiçbiri yoksa 'skipped' döner.
        """
        # 1) results_data yoksa dosyadan bulmayı dene
        if results_data is None:
            results_data = self._load_results_fallback()

        if not results_data:
            return {
                "status": "skipped",
                "reason": "No results_data provided and no cached results found."
            }

        # 2) Klasörleri hazırla
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder_name = f"{base_name}_{timestamp}"
        report_dir = Path(out_root) / folder_name / "reports"
        report_dir.mkdir(parents=True, exist_ok=True)

        # 3) Excel + JSON raporlarını üret
        excel_path = str(report_dir / "analiz_raporu.xlsx")
        json_path  = str(report_dir / "analiz_raporu.json")

        # results_data içindeki tespitlerde center/size yoksa bbox'tan türet
        normalized = self._normalize_results(results_data)

        await self.generate_excel_report(normalized, excel_path)
        await self.generate_json_report(normalized, json_path)

        return {
            "status": "ok",
            "folder_name": folder_name,
            "reports_dir": str(report_dir),
            "excel_path": excel_path,
            "json_path": json_path,
            "total_images": len(normalized),
            "total_detections": sum(len(r.get("detections", [])) for r in normalized),
        }

    # ---------------------------
    # Yardımcı: Sonuçları diskten bul
    # ---------------------------
    def _load_results_fallback(self) -> Optional[List[Dict[str, Any]]]:
        """
        results_data yoksa, standart yerlerden sonuçları okumayı dener.
        Sıra:
          1) results/analysis_results.json
          2) downloads/latest_results.json
        """
        candidates = [
            Path("results") / "analysis_results.json",
            Path("downloads") / "latest_results.json",
        ]
        for p in candidates:
            if p.exists():
                try:
                    with open(p, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    # Beklenen format: {"results":[...]} veya doğrudan [...]
                    if isinstance(data, dict) and "results" in data:
                        return data["results"]
                    if isinstance(data, list):
                        return data
                except Exception:
                    pass
        return None

    # ---------------------------
    # Yardımcı: center/size'ı garanti et
    # ---------------------------
    def _normalize_results(self, results_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Her detection için:
          - bbox -> (x1,y1,x2,y2)
          - center -> (cx,cy) yoksa bbox'tan hesaplanır
          - size   -> (w,h)   yoksa bbox'tan hesaplanır
        """
        norm = []
        for r in results_data:
            detections = []
            for d in r.get("detections", []):
                bbox = d.get("bbox") or d.get("box") or d.get("xyxy")
                if bbox and len(bbox) == 4:
                    x1, y1, x2, y2 = bbox
                    w = max(0, int(x2) - int(x1))
                    h = max(0, int(y2) - int(y1))
                    cx = int(x1 + w / 2)
                    cy = int(y1 + h / 2)
                else:
                    # bbox yoksa sıfırla
                    x1 = y1 = x2 = y2 = w = h = cx = cy = 0

                d2 = dict(d)  # kopya
                d2["bbox"] = [int(x1), int(y1), int(x2), int(y2)]
                d2.setdefault("size", [int(w), int(h)])
                d2.setdefault("center", [int(cx), int(cy)])
                # class_name yoksa class_id'den doldur
                if "class_name" not in d2:
                    cid = d2.get("class_id", 0)
                    d2["class_name"] = self.class_names.get(cid, f"Class_{cid}")
                detections.append(d2)

            norm.append({
                "filename": r.get("filename") or r.get("name") or r.get("file") or "",
                "original_path": r.get("original_path", ""),
                "processed_path": r.get("processed_path", ""),
                "detections": detections
            })
        return norm

    # ---------------------------
    # EXISTING: Excel report
    # ---------------------------
    async def generate_excel_report(self, results_data: List[Dict[str, Any]], output_path: str) -> str:
        """Generate comprehensive Excel report with multiple sheets and charts"""
        try:
            # Prepare all data
            report_data = self._prepare_detailed_data(results_data)
            summary_data = self._prepare_summary_data(results_data)
            statistics_data = self._prepare_statistics_data(results_data)
            image_analysis_data = self._prepare_image_analysis_data(results_data)

            # Create output directory if it doesn't exist
            output_dir = Path(output_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)

            # Write to Excel with multiple sheets
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 1. sheetleri yaz
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Yönetici Özeti', index=False)

                df_details = pd.DataFrame(report_data)
                df_details.to_excel(writer, sheet_name='Detaylı Sonuçlar', index=False)

                df_stats = pd.DataFrame(statistics_data)
                df_stats.to_excel(writer, sheet_name='İstatistikler', index=False)

                df_images = pd.DataFrame(image_analysis_data)
                df_images.to_excel(writer, sheet_name='Görsel Analizi', index=False)

                df_quality = pd.DataFrame(self._prepare_quality_control_data(results_data))
                df_quality.to_excel(writer, sheet_name='Kalite Kontrol', index=False)

                # flush için save() ekle
                writer.book.save(output_path)

            # with bloğu kapandıktan sonra workbook yeniden açılabilir
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            await self._format_excel_workbook(wb, {
                'Yönetici Özeti': df_summary,
                'Detaylı Sonuçlar': df_details,
                'İstatistikler': df_stats,
                'Görsel Analizi': df_images,
                'Kalite Kontrol': df_quality
            })
            wb.save(output_path)

        except Exception as e:
            print(f"Error generating Excel report: {str(e)}")
            raise

    # ---------------------------
    # EXISTING helpers (bazıları ufakça sağlamlaştırıldı)
    # ---------------------------
    def _prepare_detailed_data(self, results_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepare detailed detection data"""
        report_data = []

        for i, result in enumerate(results_data):
            filename = result.get("filename", f"image_{i}")
            detections = result.get("detections", [])

            if not detections:
                report_data.append({
                    "Fotoğraf ID": i + 1,
                    "Dosya Adı": filename,
                    "Kusur ID": "N/A",
                    "Kusur Türü": "Kusur Bulunamadı",
                    "Güven Skoru": 0.0,
                    "Güven Seviyesi": "N/A",
                    "Konum (X1, Y1, X2, Y2)": "N/A",
                    "Merkez Nokta (X, Y)": "N/A",
                    "Boyut (Genişlik x Yükseklik)": "N/A",
                    "Alan (piksel²)": 0,
                    "Kusur Yoğunluğu": "Düşük",
                    "Kritiklik Seviyesi": "Düşük",
                    "Analiz Tarihi": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Durum": "Temiz"
                })
            else:
                for j, detection in enumerate(detections):
                    bbox = detection.get('bbox', [0, 0, 0, 0])
                    center = detection.get('center')
                    size = detection.get('size')

                    # Merkez ve boyut yoksa bbox'tan hesapla
                    if not center or not size:
                        x1, y1, x2, y2 = bbox
                        w = max(0, int(x2) - int(x1))
                        h = max(0, int(y2) - int(y1))
                        cx = int(x1 + w / 2)
                        cy = int(y1 + h / 2)
                        center = [cx, cy]
                        size = [w, h]

                    area = size[0] * size[1]
                    confidence = float(detection.get("confidence", 0.0))

                    # Determine confidence level
                    if confidence >= 0.9:
                        conf_level = "Çok Yüksek"
                    elif confidence >= 0.8:
                        conf_level = "Yüksek"
                    elif confidence >= 0.7:
                        conf_level = "Orta"
                    else:
                        conf_level = "Düşük"

                    # Determine criticality based on defect type and size
                    cls_name = detection.get("class_name") or self.class_names.get(detection.get("class_id", 0), "Bilinmeyen")
                    criticality = self._determine_criticality(cls_name, area, confidence)

                    # Determine defect density
                    density = "Yüksek" if len(detections) > 3 else "Orta" if len(detections) > 1 else "Düşük"

                    report_data.append({
                        "Fotoğraf ID": i + 1,
                        "Dosya Adı": filename,
                        "Kusur ID": j + 1,
                        "Kusur Türü": cls_name,
                        "Güven Skoru": round(confidence, 3),
                        "Güven Seviyesi": conf_level,
                        "Konum (X1, Y1, X2, Y2)": f"({bbox[0]}, {bbox[1]}, {bbox[2]}, {bbox[3]})",
                        "Merkez Nokta (X, Y)": f"({center[0]}, {center[1]})",
                        "Boyut (Genişlik x Yükseklik)": f"{size[0]} x {size[1]}",
                        "Alan (piksel²)": int(area),
                        "Kusur Yoğunluğu": density,
                        "Kritiklik Seviyesi": criticality,
                        "Analiz Tarihi": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Durum": "Kusurlu"
                    })

        return report_data

    def _prepare_summary_data(self, results_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepare executive summary data"""
        total_images = len(results_data)
        total_detections = sum(len(result.get("detections", [])) for result in results_data)

        # Count by class
        class_counts = {name: 0 for name in self.class_names.values()}
        confidence_scores = []
        areas = []

        for result in results_data:
            for detection in result.get("detections", []):
                class_name = detection.get("class_name") or self.class_names.get(detection.get("class_id", 0), "Bilinmeyen")
                if class_name in class_counts:
                    class_counts[class_name] += 1
                confidence_scores.append(float(detection.get("confidence", 0.0)))
                # Alanı bbox'tan veya size'dan hesapla
                area = 0
                if 'size' in detection and detection['size']:
                    area = detection['size'][0] * detection['size'][1]
                elif 'bbox' in detection and detection['bbox']:
                    x1, y1, x2, y2 = detection['bbox']
                    area = max(0, int(x2) - int(x1)) * max(0, int(y2) - int(y1))
                areas.append(area)

        # Calculate statistics
        avg_confidence = sum(confidence_scores) / len(confidence_scores) if confidence_scores else 0
        avg_detections_per_image = total_detections / max(total_images, 1)
        avg_area = sum(areas) / len(areas) if areas else 0

        # Quality assessment
        images_with_defects = sum(1 for result in results_data if result.get("detections"))
        defect_rate = (images_with_defects / max(total_images, 1)) * 100

        quality_status = "İyi" if defect_rate < 20 else "Orta" if defect_rate < 50 else "Kötü"

        summary_data = [
            {"Metrik": "📊 GENEL BİLGİLER", "Değer": "", "Açıklama": ""},
            {"Metrik": "Toplam Fotoğraf Sayısı", "Değer": total_images, "Açıklama": "Analiz edilen toplam görsel sayısı"},
            {"Metrik": "Toplam Kusur Sayısı", "Değer": total_detections, "Açıklama": "Tespit edilen toplam kusur sayısı"},
            {"Metrik": "Kusurlu Fotoğraf Sayısı", "Değer": images_with_defects, "Açıklama": "En az bir kusur içeren fotoğraf sayısı"},
            {"Metrik": "Temiz Fotoğraf Sayısı", "Değer": total_images - images_with_defects, "Açıklama": "Hiç kusur içermeyen fotoğraf sayısı"},
            {"Metrik": "", "Değer": "", "Açıklama": ""},

            {"Metrik": "📈 İSTATİSTİKLER", "Değer": "", "Açıklama": ""},
            {"Metrik": "Ortalama Kusur/Fotoğraf", "Değer": round(avg_detections_per_image, 2), "Açıklama": "Fotoğraf başına düşen ortalama kusur sayısı"},
            {"Metrik": "Ortalama Güven Skoru", "Değer": f"{round(avg_confidence * 100, 1)}%", "Açıklama": "Tespitlerin ortalama güvenilirlik oranı"},
            {"Metrik": "Kusur Oranı", "Değer": f"{round(defect_rate, 1)}%", "Açıklama": "Kusurlu fotoğrafların toplam içindeki oranı"},
            {"Metrik": "Ortalama Kusur Alanı", "Değer": f"{round(avg_area)} piksel²", "Açıklama": "Tespit edilen kusurların ortalama alanı"},
            {"Metrik": "", "Değer": "", "Açıklama": ""},

            {"Metrik": "🎯 KUSUR TÜRÜ DAĞILIMI", "Değer": "", "Açıklama": ""},
        ]

        for class_name, count in class_counts.items():
            percentage = round((count / max(total_detections, 1)) * 100, 1)
            summary_data.append({
                "Metrik": f"  {class_name}",
                "Değer": f"{count} adet (%{percentage})",
                "Açıklama": f"Toplam {class_name.lower()} kusuru sayısı ve oranı"
            })

        summary_data.extend([
            {"Metrik": "", "Değer": "", "Açıklama": ""},
            {"Metrik": "🏆 KALİTE DEĞERLENDİRMESİ", "Değer": "", "Açıklama": ""},
            {"Metrik": "Genel Kalite Durumu", "Değer": quality_status, "Açıklama": "Kusur oranına göre genel kalite değerlendirmesi"},
            {"Metrik": "Analiz Tarihi", "Değer": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Açıklama": "Raporun oluşturulma tarihi ve saati"},
        ])

        return summary_data

    def _prepare_statistics_data(self, results_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepare detailed statistics data"""
        stats_data = []

        # Confidence distribution
        confidence_ranges = {"0.5-0.6": 0, "0.6-0.7": 0, "0.7-0.8": 0, "0.8-0.9": 0, "0.9-1.0": 0}

        # Size distribution
        size_ranges = {"Küçük (0-1000)": 0, "Orta (1000-5000)": 0, "Büyük (5000+)": 0}

        # Process all detections
        all_detections = []
        for result in results_data:
            all_detections.extend(result.get("detections", []))

        for detection in all_detections:
            confidence = float(detection.get("confidence", 0.0))
            # Alan
            if 'size' in detection and detection['size']:
                area = detection['size'][0] * detection['size'][1]
            elif 'bbox' in detection and detection['bbox']:
                x1, y1, x2, y2 = detection['bbox']
                area = max(0, int(x2) - int(x1)) * max(0, int(y2) - int(y1))
            else:
                area = 0

            # Confidence distribution
            if 0.5 <= confidence < 0.6:
                confidence_ranges["0.5-0.6"] += 1
            elif 0.6 <= confidence < 0.7:
                confidence_ranges["0.6-0.7"] += 1
            elif 0.7 <= confidence < 0.8:
                confidence_ranges["0.7-0.8"] += 1
            elif 0.8 <= confidence < 0.9:
                confidence_ranges["0.8-0.9"] += 1
            elif 0.9 <= confidence <= 1.0:
                confidence_ranges["0.9-1.0"] += 1

            # Size distribution
            if area < 1000:
                size_ranges["Küçük (0-1000)"] += 1
            elif area < 5000:
                size_ranges["Orta (1000-5000)"] += 1
            else:
                size_ranges["Büyük (5000+)"] += 1

        # Add confidence statistics
        stats_data.append({"Kategori": "GÜVENİLİRLİK DAĞILIMI", "Alt Kategori": "", "Sayı": "", "Yüzde": ""})
        total_detections = len(all_detections)
        for range_name, count in confidence_ranges.items():
            percentage = (count / max(total_detections, 1)) * 100
            stats_data.append({
                "Kategori": "Güven Aralığı",
                "Alt Kategori": range_name,
                "Sayı": count,
                "Yüzde": f"{percentage:.1f}%"
            })

        # Add size statistics
        stats_data.append({"Kategori": "", "Alt Kategori": "", "Sayı": "", "Yüzde": ""})
        stats_data.append({"Kategori": "BOYUT DAĞILIMI", "Alt Kategori": "", "Sayı": "", "Yüzde": ""})
        for range_name, count in size_ranges.items():
            percentage = (count / max(total_detections, 1)) * 100
            stats_data.append({
                "Kategori": "Kusur Boyutu",
                "Alt Kategori": range_name,
                "Sayı": count,
                "Yüzde": f"{percentage:.1f}%"
            })

        return stats_data

    def _prepare_image_analysis_data(self, results_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepare per-image analysis data"""
        image_data = []

        for i, result in enumerate(results_data):
            filename = result.get("filename", f"image_{i}")
            detections = result.get("detections", [])

            # Calculate image-level statistics
            total_defects = len(detections)
            avg_confidence = (sum(float(d.get("confidence", 0.0)) for d in detections) / max(total_defects, 1)) if detections else 0

            # Count by type
            type_counts = {name: 0 for name in self.class_names.values()}
            total_area = 0

            for detection in detections:
                class_name = detection.get("class_name") or self.class_names.get(detection.get("class_id", 0), "Bilinmeyen")
                if class_name in type_counts:
                    type_counts[class_name] += 1

                if 'size' in detection and detection['size']:
                    total_area += detection['size'][0] * detection['size'][1]
                elif 'bbox' in detection and detection['bbox']:
                    x1, y1, x2, y2 = detection['bbox']
                    total_area += max(0, int(x2) - int(x1)) * max(0, int(y2) - int(y1))

            # Determine image quality
            if total_defects == 0:
                quality = "Mükemmel"
                risk_level = "Düşük"
            elif total_defects <= 2 and avg_confidence < 0.8:
                quality = "İyi"
                risk_level = "Düşük"
            elif total_defects <= 5:
                quality = "Orta"
                risk_level = "Orta"
            else:
                quality = "Kötü"
                risk_level = "Yüksek"

            image_data.append({
                "Fotoğraf Adı": filename,
                "Toplam Kusur": total_defects,
                "Krater": type_counts.get("Krater", 0),
                "Tanecik": type_counts.get("Tanecik", 0),
                "Pinhol": type_counts.get("Pinhol", 0),
                "Ortalama Güven": f"{avg_confidence * 100:.1f}%" if avg_confidence > 0 else "N/A",
                "Toplam Kusur Alanı": f"{int(total_area)} piksel²",
                "Kalite Değerlendirmesi": quality,
                "Risk Seviyesi": risk_level,
                "Öneri": self._get_recommendation(total_defects, type_counts, avg_confidence)
            })

        return image_data

    def _prepare_quality_control_data(self, results_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepare quality control checklist data"""
        total_images = len(results_data)
        total_detections = sum(len(result.get("detections", [])) for result in results_data)

        # Quality thresholds
        max_defects_per_image = 3
        min_confidence_threshold = 0.7
        max_defect_rate = 0.3  # 30%

        # Calculate metrics
        images_with_many_defects = sum(1 for result in results_data if len(result.get("detections", [])) > max_defects_per_image)
        low_confidence_detections = 0

        for result in results_data:
            for detection in result.get("detections", []):
                if float(detection.get("confidence", 0.0)) < min_confidence_threshold:
                    low_confidence_detections += 1

        defect_rate = (sum(1 for result in results_data if result.get("detections")) / max(total_images, 1))

        quality_data = [
            {
                "Kontrol Noktası": "Toplam Görsel Sayısı",
                "Değer": total_images,
                "Hedef/Limit": "N/A",
                "Durum": "✓ Tamam",
                "Açıklama": "Analiz edilen toplam görsel sayısı"
            },
            {
                "Kontrol Noktası": "Yüksek Kusur Yoğunluklu Görseller",
                "Değer": images_with_many_defects,
                "Hedef/Limit": f"< {max_defects_per_image} kusur/görsel",
                "Durum": "✓ Tamam" if images_with_many_defects == 0 else "⚠ Dikkat" if images_with_many_defects < total_images * 0.1 else "✗ Problem",
                "Açıklama": f"{max_defects_per_image}'den fazla kusur içeren görsel sayısı"
            },
            {
                "Kontrol Noktası": "Düşük Güvenilirlik Tespitleri",
                "Değer": low_confidence_detections,
                "Hedef/Limit": f"> {min_confidence_threshold * 100}% güven",
                "Durum": "✓ Tamam" if low_confidence_detections == 0 else "⚠ Dikkat" if low_confidence_detections < total_detections * 0.1 else "✗ Problem",
                "Açıklama": f"{min_confidence_threshold * 100}%'den düşük güvenilirlikte tespit sayısı"
            },
            {
                "Kontrol Noktası": "Genel Kusur Oranı",
                "Değer": f"{defect_rate * 100:.1f}%",
                "Hedef/Limit": f"< {max_defect_rate * 100}%",
                "Durum": "✓ Tamam" if defect_rate < max_defect_rate else "⚠ Dikkat" if defect_rate < max_defect_rate * 1.5 else "✗ Problem",
                "Açıklama": "Kusurlu görsellerin toplam içindeki oranı"
            },
            {
                "Kontrol Noktası": "Model Performansı",
                "Değer": "Aktif",
                "Hedef/Limit": "Çalışır durumda",
                "Durum": "✓ Tamam",
                "Açıklama": "AI modelinin çalışma durumu"
            }
        ]

        return quality_data

    def _determine_criticality(self, defect_type: str, area: int, confidence: float) -> str:
        """Determine criticality level based on defect characteristics"""
        # Base criticality by type
        type_criticality = {
            "Krater": 3,      # High criticality
            "Pinhol": 2,      # Medium criticality
            "Tanecik": 1      # Low criticality
        }

        base_score = type_criticality.get(defect_type, 1)

        # Adjust by size
        if area > 5000:
            base_score += 2
        elif area > 2000:
            base_score += 1

        # Adjust by confidence
        if confidence > 0.9:
            base_score += 1

        # Determine final criticality
        if base_score >= 5:
            return "Kritik"
        elif base_score >= 3:
            return "Yüksek"
        elif base_score >= 2:
            return "Orta"
        else:
            return "Düşük"

    def _get_recommendation(self, total_defects: int, type_counts: Dict[str, int], avg_confidence: float) -> str:
        """Get recommendation based on analysis results"""
        if total_defects == 0:
            return "Kalite standartlarına uygun. Üretim devam edebilir."
        elif total_defects <= 2 and avg_confidence < 0.8:
            return "Minör kusurlar mevcut. Kontrol edilmesi önerilir."
        elif type_counts.get("Krater", 0) > 0:
            return "Krater kusuru tespit edildi. Acil müdahale gerekli."
        elif total_defects > 5:
            return "Yüksek kusur yoğunluğu. Üretim parametreleri gözden geçirilmeli."
        else:
            return "Orta seviye kusurlar. Kalite kontrol süreçleri iyileştirilebilir."

    async def _format_excel_workbook(self, workbook, dataframes: Dict[str, pd.DataFrame]):
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # Define styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            subheader_font = Font(bold=True, size=11, color="2F4F4F")
            subheader_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Format each sheet
            for sheet_name, df in dataframes.items():
                worksheet = workbook[sheet_name]

                # Format headers
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

                # Add borders to all cells
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = border

                # Special formatting for summary sheet
                if sheet_name == 'Yönetici Özeti':
                    for row in worksheet.iter_rows(min_row=2):
                        if row[0].value and str(row[0].value).startswith(('📊', '📈', '🎯', '🏆')):
                            for cell in row:
                                cell.font = subheader_font
                                cell.fill = subheader_fill

            # Chart ekleme
            if 'İstatistikler' in dataframes:
                await self._add_charts_to_statistics(workbook['İstatistikler'])

        except Exception as e:
            print(f"Warning: Could not apply advanced formatting: {str(e)}")


    async def _add_charts_to_statistics(self, worksheet):
        """Add charts to statistics sheet"""
        try:
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.chart.label import DataLabelList

            # Basit bir chart alanı hazırlığı (veri konumları projeye göre özelleştirilebilir)
            chart = BarChart()
            chart.title = "Güvenilirlik Dağılımı"
            chart.x_axis.title = "Güven Aralığı"
            chart.y_axis.title = "Tespit Sayısı"
            worksheet.add_chart(chart, "F2")

        except Exception as e:
            print(f"Warning: Could not add charts: {str(e)}")

    async def generate_pdf_report(self, results_data: List[Dict[str, Any]], output_path: str) -> str:
        """Generate PDF report as an alternative to Excel"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors

            # Create PDF document
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph("Boya Kusurları Analiz Raporu", title_style))
            story.append(Spacer(1, 20))

            # Summary section
            summary_data = self._prepare_summary_data(results_data)
            story.append(Paragraph("Yönetici Özeti", styles['Heading2']))

            # Create summary table
            summary_table_data = [["Metrik", "Değer", "Açıklama"]]
            for item in summary_data[:10]:  # First 10 items
                if item["Metrik"] and not item["Metrik"].startswith(('📊', '📈', '🎯', '🏆')):
                    summary_table_data.append([
                        item["Metrik"],
                        str(item["Değer"]),
                        item["Açıklama"]
                    ])

            from reportlab.platypus import Table  # local import safety
            summary_table = Table(summary_table_data, colWidths=[2*inch, 1*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            story.append(summary_table)
            story.append(PageBreak())

            # Detailed results section
            story.append(Paragraph("Detaylı Sonuçlar", styles['Heading2']))

            detailed_data = self._prepare_detailed_data(results_data)
            if detailed_data:
                # Create detailed table (first 20 entries)
                detail_table_data = [["Dosya", "Kusur Türü", "Güven", "Kritiklik"]]
                for item in detailed_data[:20]:
                    detail_table_data.append([
                        item["Dosya Adı"],
                        item["Kusur Türü"],
                        f"{item['Güven Skoru']:.2f}",
                        item["Kritiklik Seviyesi"]
                    ])

                detail_table = Table(detail_table_data, colWidths=[2*inch, 1.5*inch, 1*inch, 1.5*inch])
                detail_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))

                story.append(detail_table)

            # Build PDF
            doc.build(story)

            print(f"PDF report generated: {output_path}")
            return output_path

        except ImportError:
            print("Warning: PDF generation requires reportlab package")
            raise RuntimeError("PDF generation not available - install reportlab package")
        except Exception as e:
            print(f"Error generating PDF report: {str(e)}")
            raise

    async def generate_json_report(self, results_data: List[Dict[str, Any]], output_path: str) -> str:
        """Generate JSON report from analysis results"""
        try:
            # Prepare comprehensive report structure
            report = {
                "metadata": {
                    "generated_at": datetime.now().isoformat(),
                    "version": "2.0.0",
                    "report_type": "paint_defect_analysis",
                    "total_images": len(results_data),
                    "total_detections": sum(len(result.get("detections", [])) for result in results_data)
                },
                "executive_summary": self._generate_summary_stats(results_data),
                "detailed_results": results_data,
                "statistics": {
                    "confidence_distribution": self._calculate_confidence_distribution(results_data),
                    "size_distribution": self._calculate_size_distribution(results_data),
                    "defect_type_analysis": self._calculate_defect_type_analysis(results_data)
                },
                "quality_assessment": self._generate_quality_assessment(results_data),
                "recommendations": self._generate_recommendations(results_data)
            }

            # Create output directory if it doesn't exist
            output_dir = Path(output_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)

            # Write JSON file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(report, f, ensure_ascii=False, indent=2)

            print(f"JSON report generated: {output_path}")
            return output_path

        except Exception as e:
            print(f"Error generating JSON report: {str(e)}")
            raise

    def _calculate_confidence_distribution(self, results_data: List[Dict[str, Any]]) -> Dict[str, int]:
        """Calculate confidence score distribution"""
        distribution = {"0.5-0.6": 0, "0.6-0.7": 0, "0.7-0.8": 0, "0.8-0.9": 0, "0.9-1.0": 0}

        for result in results_data:
            for detection in result.get("detections", []):
                confidence = float(detection.get("confidence", 0.0))
                if 0.5 <= confidence < 0.6:
                    distribution["0.5-0.6"] += 1
                elif 0.6 <= confidence < 0.7:
                    distribution["0.6-0.7"] += 1
                elif 0.7 <= confidence < 0.8:
                    distribution["0.7-0.8"] += 1
                elif 0.8 <= confidence < 0.9:
                    distribution["0.8-0.9"] += 1
                elif 0.9 <= confidence <= 1.0:
                    distribution["0.9-1.0"] += 1

        return distribution

    def _calculate_size_distribution(self, results_data: List[Dict[str, Any]]) -> Dict[str, int]:
        """Calculate defect size distribution"""
        distribution = {"small": 0, "medium": 0, "large": 0}

        for result in results_data:
            for detection in result.get("detections", []):
                area = 0
                if 'size' in detection and detection['size']:
                    area = detection['size'][0] * detection['size'][1]
                elif 'bbox' in detection and detection['bbox']:
                    x1, y1, x2, y2 = detection['bbox']
                    area = max(0, int(x2) - int(x1)) * max(0, int(y2) - int(y1))

                if area < 1000:
                    distribution["small"] += 1
                elif area < 5000:
                    distribution["medium"] += 1
                else:
                    distribution["large"] += 1

        return distribution

    def _calculate_defect_type_analysis(self, results_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate detailed defect type analysis"""
        analysis = {}

        for class_name in self.class_names.values():
            analysis[class_name] = {
                "count": 0,
                "avg_confidence": 0.0,
                "avg_size": 0.0,
                "criticality_distribution": {"Düşük": 0, "Orta": 0, "Yüksek": 0, "Kritik": 0}
            }

        # Collect data
        for result in results_data:
            for detection in result.get("detections", []):
                cls = detection.get("class_name") or self.class_names.get(detection.get("class_id", 0), "Bilinmeyen")
                rec = analysis.setdefault(cls, {
                    "count": 0, "avg_confidence": 0.0, "avg_size": 0.0,
                    "criticality_distribution": {"Düşük": 0, "Orta": 0, "Yüksek": 0, "Kritik": 0}
                })
                rec["count"] += 1

                # Add confidence
                rec["avg_confidence"] += float(detection.get("confidence", 0.0))

                # Add size
                area = 0
                if 'size' in detection and detection['size']:
                    area = detection['size'][0] * detection['size'][1]
                elif 'bbox' in detection and detection['bbox']:
                    x1, y1, x2, y2 = detection['bbox']
                    area = max(0, int(x2) - int(x1)) * max(0, int(y2) - int(y1))
                rec["avg_size"] += float(area)

                # Add criticality
                criticality = self._determine_criticality(
                    cls,
                    area,
                    float(detection.get("confidence", 0.0))
                )
                rec["criticality_distribution"][criticality] = rec["criticality_distribution"].get(criticality, 0) + 1

        # Calculate averages
        for cls, rec in analysis.items():
            count = rec["count"]
            if count > 0:
                rec["avg_confidence"] = rec["avg_confidence"] / count
                rec["avg_size"] = rec["avg_size"] / count
            else:
                rec["avg_confidence"] = 0.0
                rec["avg_size"] = 0.0

        return analysis

    def _generate_quality_assessment(self, results_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Generate overall quality assessment"""
        total_images = len(results_data)
        images_with_defects = sum(1 for result in results_data if result.get("detections"))
        defect_rate = images_with_defects / max(total_images, 1)

        # Determine quality grade
        if defect_rate < 0.1:
            grade = "A"
            status = "Mükemmel"
        elif defect_rate < 0.3:
            grade = "B"
            status = "İyi"
        elif defect_rate < 0.5:
            grade = "C"
            status = "Orta"
        else:
            grade = "D"
            status = "Kötü"

        return {
            "overall_grade": grade,
            "quality_status": status,
            "defect_rate": round(defect_rate * 100, 1),
            "pass_rate": round((1 - defect_rate) * 100, 1),
            "total_images_analyzed": total_images,
            "images_with_defects": images_with_defects,
            "clean_images": total_images - images_with_defects
        }

    def _generate_recommendations(self, results_data: List[Dict[str, Any]]) -> List[str]:
        """Generate actionable recommendations"""
        recommendations = []

        # Analyze patterns
        total_detections = sum(len(result.get("detections", [])) for result in results_data)
        class_counts = {name: 0 for name in self.class_names.values()}

        for result in results_data:
            for detection in result.get("detections", []):
                class_name = detection.get("class_name") or self.class_names.get(detection.get("class_id", 0), "Bilinmeyen")
                if class_name in class_counts:
                    class_counts[class_name] += 1

        # Generate specific recommendations
        if total_detections > 0:
            if class_counts.get("Krater", 0) > total_detections * 0.3:
                recommendations.append("Krater kusurları yüksek oranda tespit edildi. Yüzey hazırlama süreçlerini gözden geçirin.")

            if class_counts.get("Pinhol", 0) > total_detections * 0.3:
                recommendations.append("Pinhol kusurları yaygın. Boya viskozitesi ve uygulama basıncını kontrol edin.")

            if class_counts.get("Tanecik", 0) > total_detections * 0.3:
                recommendations.append("Tanecik kusurları mevcut. Filtrasyon sistemlerini ve temizlik prosedürlerini iyileştirin.")

        # General recommendations
        defect_rate = sum(1 for result in results_data if result.get("detections")) / max(len(results_data), 1)

        if defect_rate > 0.5:
            recommendations.append("Yüksek kusur oranı tespit edildi. Üretim parametrelerinin kapsamlı incelenmesi önerilir.")
        elif defect_rate > 0.3:
            recommendations.append("Orta seviye kusur oranı. Kalite kontrol süreçlerinin sıklaştırılması önerilir.")

        if not recommendations:
            recommendations.append("Kalite standartları genel olarak iyi durumda. Mevcut süreçlerin devamı önerilir.")

        return recommendations

    def _generate_summary_stats(self, results_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Generate summary statistics"""
        total_images = len(results_data)
        total_detections = sum(len(result.get("detections", [])) for result in results_data)

        # Count by class
        class_counts = {name: 0 for name in self.class_names.values()}
        confidence_scores = []

        for result in results_data:
            for detection in result.get("detections", []):
                class_name = detection.get("class_name") or self.class_names.get(detection.get("class_id", 0), "Bilinmeyen")
                if class_name in class_counts:
                    class_counts[class_name] += 1
                confidence_scores.append(float(detection.get("confidence", 0.0)))

        # Calculate statistics
        avg_confidence = sum(confidence_scores) / len(confidence_scores) if confidence_scores else 0.0
        avg_detections_per_image = total_detections / max(total_images, 1)

        return {
            "total_images": total_images,
            "total_detections": total_detections,
            "average_detections_per_image": round(avg_detections_per_image, 2),
            "average_confidence": round(avg_confidence, 3),
            "class_distribution": class_counts,
            "images_with_defects": sum(1 for result in results_data if result.get("detections")),
            "images_without_defects": sum(1 for result in results_data if not result.get("detections"))
        }
